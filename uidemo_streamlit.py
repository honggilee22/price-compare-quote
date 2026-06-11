# -*- coding: utf-8 -*-
from __future__ import annotations

import base64
from datetime import date, datetime
from email.message import EmailMessage
from io import BytesIO
import hashlib
import json
import os
from pathlib import Path
import platform
import re
import shutil
import smtplib
import subprocess
import sys
import tempfile
import time

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment
from pypdf import PdfReader, PdfWriter
import streamlit as st
import streamlit.components.v1 as components

HTML_PATH = Path(__file__).with_name("uidemo.html")
TEMPLATE_PATH = Path(__file__).with_name("SS.xlsx")
SINGLE_TEMPLATE_PATH = Path(__file__).with_name("SS (1).xlsx")
COMPONENT_DIR = Path(__file__).with_name("components") / "price_compare"
CONFIG_PATH = Path(__file__).with_name("output") / "app_settings.json"
DEFAULT_SAVE_DIR = Path(__file__).with_name("output") / "quotes"
DEBUG_LOG_PATH = Path(__file__).with_name("output") / "pdf_debug.log"
PDF_WORK_DIR = Path(__file__).with_name("output") / "tmp_pdf_work"
CATALOG_PATH = Path(__file__).with_name("0817_with_air_no_links_fixed.html")
DEMO1_PATH = Path(__file__).with_name("demo1.html")
INTRO_PDF_SOURCE = Path(__file__).with_name("생활가전_렌탈_제안서.pdf")
EXCEL_PDF_WORKER = Path(__file__).with_name("excel_pdf_worker.py")
CATALOG_LINKS = [
    ("슈퍼아이스트리", "https://planner-power.com/catalog/prod_detail.php?idx=104&pp=2&rp=16"),
    ("700", "https://planner-power.com/catalog/prod_detail.php?idx=110&pp=2&rp=16"),
    ("550", "https://planner-power.com/catalog/prod_detail.php?idx=71&pp=2&rp=16"),
    ("토타max", "https://planner-power.com/catalog/prod_detail.php?idx=165&pp=1&rp=8"),
    ("토타", "https://planner-power.com/catalog/prod_detail.php?idx=108&pp=1&rp=12"),
    ("서밋", "https://planner-power.com/catalog/prod_detail.php?idx=173&pp=2&rp=24"),
    ("디오", "https://planner-power.com/catalog/prod_detail.php?idx=139&pp=1&rp=14"),
    ("B400", "https://planner-power.com/catalog/prod_detail.php?idx=164&pp=1&rp=12"),
    ("B350", "https://planner-power.com/catalog/prod_detail.php?idx=95&pp=2&rp=24"),
    ("루컷", "https://planner-power.com/catalog/prod_detail.php?idx=140&pp=1&rp=2"),
]
CHUNGHO_PRODUCT_LINKS = dict(CATALOG_LINKS)
DEFAULT_REMARKS = "@상품명 클릭하시면 카달로그 페이지로 이동 합니다"
HARDCODE_GMAIL_USER = ""
HARDCODE_GMAIL_APP_PASSWORD = ""

MAX_TEMPLATE_ROWS = 3
COMPONENT_HEIGHT = 1200

TEMPLATE_CONFIGS = {
    "compare": {
        "print_area": "A1:T30",
        "recipient_cell": "B4",
        "summary_cell": "B13",
        "date_cell": "D5",
        "ext_cell": "D6",
        "email_cell": "D7",
        "plan1_total_cell": "E11",
        "plan2_total_cell": "N11",
        "plan1_prepay_cell": "E12",
        "plan2_prepay_cell": "N12",
        "plan1_rows": [15, 17, 19],
        "plan1_cols": {"model": "B", "price": "E", "qty": "G", "total": "H"},
        "plan2_rows": [15, 17, 19],
        "plan2_cols": {"model": "K", "price": "N", "qty": "P", "total": "Q"},
        "plan2_promo_offset": 1,
        "remarks_cell": "B25",
        "clear_compare_rows_when_single": False,
    },
    "single": {
        "print_area": "A1:T30",
        "recipient_cell": "B4",
        "summary_cell": None,
        "date_cell": "D5",
        "ext_cell": "D6",
        "email_cell": "D7",
        "plan1_total_cell": None,
        "plan2_total_cell": None,
        "plan1_prepay_cell": None,
        "plan2_prepay_cell": None,
        "plan1_rows": [12, 14, 16],
        "plan1_cols": {"model": "B", "price": "G", "qty": "K", "total": "N"},
        "plan2_rows": [],
        "plan2_cols": {},
        "plan2_promo_offset": 0,
        "remarks_cell": "B24",
        "clear_compare_rows_when_single": False,
    },
}

BRIDGE_SCRIPT = """<script id=\"price-compare-bridge\">
(function() {
  function buildPayload(action) {
    const recipient = (document.getElementById("quote-recipient")?.value || "").trim();
    const ext = (document.getElementById("quote-ext")?.value || "").trim();
    const quoteDate = document.getElementById("quote-date")?.value || "";
    const email = (document.getElementById("quote-email")?.value || "").trim();
    const remarks = (document.getElementById("quote-remarks")?.value || "").trim();
    const mode = document.querySelector(".mode-switch button.active")?.dataset.mode || "compare";

    const plan1El = document.querySelector('[data-plan="1"]');
    const plan2El = document.querySelector('[data-plan="2"]');

    const plan1 = collectPlanData(plan1El, "1안");
    const plan2 = collectPlanData(plan2El, "2안");

    const discount1 = getDiscount(plan1El) || 0;
    const discount2 = getDiscount(plan2El) || 0;

    return {
      request_id: String(Date.now()),
      action: action,
      data: {
        recipient: recipient,
        ext: ext,
        quote_date: quoteDate,
        email: email,
        remarks: remarks,
        mode: mode,
        plan1: plan1,
        plan2: plan2,
        discount1: discount1,
        discount2: discount2
      }
    };
  }

  function post(action) {
    const payload = buildPayload(action);
    window.parent.postMessage({ source: "price-compare", payload: payload }, "*");
  }

  window.PriceCompareBridge = {
    post: post,
    buildPayload: buildPayload
  };

  function init() {}

  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", init);
  } else {
    init();
  }
})();
</script>
"""

RESPONSIVE_CSS = """
.container{width:100%;max-width:960px;margin:0 auto;}
img,video{max-width:100%;height:auto;}
body{overflow-x:hidden;}
.canvas{width:100%;max-width:960px;margin:0 auto;height:auto;min-height:100vh;}
@media (max-width:768px){.container{padding:0 16px;}.canvas{padding:0 16px;}}
"""


def append_debug_log(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {message}"
    print(line, flush=True)
    try:
        DEBUG_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with DEBUG_LOG_PATH.open("a", encoding="utf-8") as log_file:
            log_file.write(f"{line}\n")
    except Exception:
        pass


def load_settings():
    if not CONFIG_PATH.exists():
        return {}
    try:
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def persist_save_dir(raw_value):
    save_path = Path(raw_value).expanduser()
    if not save_path.is_absolute():
        save_path = Path(__file__).parent / save_path
    save_path.mkdir(parents=True, exist_ok=True)
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {"save_dir": str(save_path)}
    CONFIG_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return payload["save_dir"]


def inject_bridge(html_text):
    if "price-compare-bridge" in html_text:
        return html_text
    if "</body>" in html_text:
        return html_text.replace("</body>", f"{BRIDGE_SCRIPT}\n</body>")
    return f"{html_text}\n{BRIDGE_SCRIPT}"


def inject_responsive_layout(html_text):
    updated = html_text
    viewport_tag = (
        '<meta name="viewport" content="width=device-width, initial-scale=1.0">'
    )
    if re.search(r'<meta name="viewport"[^>]*>', updated, flags=re.IGNORECASE):
        updated = re.sub(
            r'<meta name="viewport"[^>]*>',
            viewport_tag,
            updated,
            flags=re.IGNORECASE,
        )
    elif "</head>" in updated:
        updated = updated.replace("</head>", f"{viewport_tag}\n</head>")
    else:
        updated = f"{viewport_tag}\n{updated}"

    if "responsive-layout" in updated:
        return updated
    style_tag = f'<style id="responsive-layout">{RESPONSIVE_CSS}</style>'
    if "</head>" in updated:
        return updated.replace("</head>", f"{style_tag}\n</head>")
    return f"{style_tag}\n{updated}"


def read_text_flexible(path):
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="cp949", errors="replace")


def extract_catalog_data(html_text):
    marker = "const data ="
    marker_index = html_text.find(marker)
    if marker_index == -1:
        return None
    start_index = html_text.find("{", marker_index)
    if start_index == -1:
        return None
    depth = 0
    for index in range(start_index, len(html_text)):
        char = html_text[index]
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                return html_text[start_index : index + 1]
    return None


def parse_catalog_json(data_text):
    try:
        return json.loads(data_text)
    except json.JSONDecodeError:
        return None


def inject_catalog_data(html_text):
    if "window.catalogData =" in html_text or "window.catalogData=" in html_text:
        return html_text
    catalog_data = None
    html_data_text = None
    if CATALOG_PATH.exists():
        source_text = read_text_flexible(CATALOG_PATH)
        html_data_text = extract_catalog_data(source_text)
        if html_data_text:
            catalog_data = parse_catalog_json(html_data_text)
    if catalog_data is not None:
        data_text = json.dumps(catalog_data, ensure_ascii=False)
    else:
        return html_text
    script = f"<script id=\"catalog-data\">window.catalogData = {data_text};</script>"
    if "<body>" in html_text:
        return html_text.replace("<body>", f"<body>\n{script}", 1)
    if "</head>" in html_text:
        return html_text.replace("</head>", f"{script}\n</head>", 1)
    return f"{script}\n{html_text}"


def inject_product_info(html_text):
    if "product-info" in html_text:
        return html_text
    if not DEMO1_PATH.exists():
        return html_text
    product_html = read_text_flexible(DEMO1_PATH)
    encoded = base64.b64encode(product_html.encode("utf-8")).decode("ascii")
    script = (
        "<script id=\"product-info\">"
        "(function(){"
        f"const raw=\"{encoded}\";"
        "try{"
        "const bytes=Uint8Array.from(atob(raw),c=>c.charCodeAt(0));"
        "window.productInfoHtml=new TextDecoder('utf-8').decode(bytes);"
        "}catch(e){window.productInfoHtml='';}"
        "})();"
        "</script>"
    )
    if "<body>" in html_text:
        return html_text.replace("<body>", f"<body>\n{script}", 1)
    if "</head>" in html_text:
        return html_text.replace("</head>", f"{script}\n</head>", 1)
    return f"{script}\n{html_text}"


def inject_initial_view(html_text, initial_view):
    if "initial-view" in html_text:
        return html_text
    view = (initial_view or "price").strip() or "price"
    script = f"<script id=\"initial-view\">window.initialView=\"{view}\";</script>"
    if "<body>" in html_text:
        return html_text.replace("<body>", f"<body>\n{script}", 1)
    if "</head>" in html_text:
        return html_text.replace("</head>", f"{script}\n</head>", 1)
    return f"{script}\n{html_text}"


def sanitize_filename(text):
    cleaned = re.sub(r"[\\/:*?\"<>|]", "_", text or "")
    cleaned = re.sub(r"\s+", "", cleaned)
    return cleaned or "견적서"


def format_date_label(quote_date):
    if not quote_date:
        return ""
    return quote_date.strftime("%Y.%m.%d")


def format_file_date(quote_date):
    if not quote_date:
        return date.today().strftime("%m월%d일")
    return quote_date.strftime("%m월%d일")


def parse_date(value):
    if not value:
        return date.today()
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return date.today()


def parse_number(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return 0
    text = re.sub(r"[^\d]", "", text)
    return int(text or 0)


def parse_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def normalize_rows(rows):
    normalized = []
    for row in rows:
        model = str(row.get("model", "") or "").strip()
        price = parse_number(row.get("price"))
        qty = parse_number(row.get("qty"))
        promo_price = parse_number(row.get("promo_price"))
        if model or price or qty or promo_price:
            normalized.append(
                {"model": model, "price": price, "qty": qty, "promo_price": promo_price}
            )
    if not normalized:
        normalized = [{"model": "", "price": 0, "qty": 0, "promo_price": 0}]
    while len(normalized) < MAX_TEMPLATE_ROWS:
        normalized.append({"model": "", "price": 0, "qty": 0, "promo_price": 0})
    return normalized[:MAX_TEMPLATE_ROWS]


def compute_totals(rows, use_promo=False):
    total_sum = 0
    for row in rows:
        price = parse_number(row.get("price"))
        qty = parse_number(row.get("qty"))
        promo_price = parse_number(row.get("promo_price"))
        effective_price = promo_price if use_promo and promo_price else price
        total_sum += effective_price * qty
    return total_sum


def format_won(value):
    try:
        amount = int(round(float(value)))
    except (TypeError, ValueError):
        amount = 0
    return f"{amount:,}원"


def build_summary_text(total1, total2):
    if total2 < total1:
        diff = total1 - total2
        return f"연 {format_won(diff * 12)} 할인 / 월 {format_won(diff)} 할인"
    if total2 > total1:
        diff = total2 - total1
        return f"월 {format_won(diff)} 추가"
    return "차이 0원"


def load_template(mode):
    template_path = SINGLE_TEMPLATE_PATH if (mode or "").strip() == "single" else TEMPLATE_PATH
    if not template_path.exists():
        raise FileNotFoundError(f"템플릿을 찾을 수 없습니다: {template_path}")
    return openpyxl.load_workbook(template_path)


def resolve_target_cell(worksheet, cell_ref):
    cell = worksheet[cell_ref]
    if cell.__class__.__name__ != "MergedCell":
        return cell
    row = cell.row
    col = cell.column
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return worksheet.cell(merged_range.min_row, merged_range.min_col)
    return cell


def write_cell(worksheet, cell_ref, value):
    resolve_target_cell(worksheet, cell_ref).value = value


def write_model_cell(worksheet, cell_ref, model):
    cell = resolve_target_cell(worksheet, cell_ref)
    model_text = str(model or "").strip()
    product_url = CHUNGHO_PRODUCT_LINKS.get(model_text)
    if product_url:
        escaped_url = product_url.replace('"', '""')
        escaped_model = model_text.replace('"', '""')
        cell.value = f'=HYPERLINK("{escaped_url}","{escaped_model}")'
        cell.font = cell.font.copy(color="0563C1", bold=True, underline="single")
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical,
            text_rotation=cell.alignment.text_rotation,
            wrap_text=False,
            shrink_to_fit=True,
            indent=cell.alignment.indent,
        )
        return
    cell.value = model_text
    cell.hyperlink = None


def write_remarks_cell(worksheet, cell_ref, remarks):
    cell = resolve_target_cell(worksheet, cell_ref)
    remarks_text = str(remarks or "").strip() or DEFAULT_REMARKS
    if remarks_text == DEFAULT_REMARKS:
        cell.value = CellRichText(
            TextBlock(InlineFont(color="0563C1", b=True), "@상품명"),
            " 클릭하시면 카달로그 페이지로 이동 합니다",
        )
        return
    cell.value = remarks_text


def get_template_config(mode):
    return TEMPLATE_CONFIGS["single" if (mode or "").strip() == "single" else "compare"]


def configure_a4_sheet(worksheet, print_area):
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = None
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.page_setup.scale = None
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0
    worksheet.page_margins.right = 0
    worksheet.page_margins.top = 0
    worksheet.page_margins.bottom = 0
    worksheet.page_margins.header = 0
    worksheet.page_margins.footer = 0
    worksheet.print_area = print_area


def prepare_quote_workbook(workbook, mode):
    quote_sheet = workbook.active
    quote_sheet._images = []
    config = get_template_config(mode)
    configure_a4_sheet(quote_sheet, config["print_area"])
    return quote_sheet


def fill_template(
    recipient,
    ext,
    quote_date,
    email,
    mode,
    plan1_rows,
    plan2_rows,
    plan1_total,
    plan2_total,
    plan1_prepay,
    plan2_prepay,
    plan2_promo_enabled=False,
    summary_text="",
    remarks="",
):
    wb = load_template(mode)
    ws = prepare_quote_workbook(wb, mode)
    is_single_mode = (mode or "").strip() == "single"
    config = get_template_config(mode)

    write_cell(ws, config["recipient_cell"], recipient or "")
    if config.get("summary_cell"):
        write_cell(ws, config["summary_cell"], summary_text or "")
    write_cell(ws, config["date_cell"], format_date_label(quote_date))
    write_cell(ws, config["ext_cell"], ext or "")
    write_cell(ws, config["email_cell"], email or "")
    write_remarks_cell(ws, config["remarks_cell"], remarks)

    if config.get("plan1_total_cell"):
        write_cell(ws, config["plan1_total_cell"], plan1_total)
    if config.get("plan2_total_cell"):
        write_cell(ws, config["plan2_total_cell"], plan2_total)
    if config.get("plan1_prepay_cell"):
        write_cell(ws, config["plan1_prepay_cell"], plan1_prepay)
    if config.get("plan2_prepay_cell"):
        write_cell(ws, config["plan2_prepay_cell"], plan2_prepay)

    for index, row in enumerate(plan1_rows[: len(config["plan1_rows"])]):
        excel_row = config["plan1_rows"][index]
        plan1_cols = config["plan1_cols"]
        model = row.get("model", "")
        price = parse_number(row.get("price"))
        qty = parse_number(row.get("qty"))
        total = price * qty
        write_model_cell(ws, f"{plan1_cols['model']}{excel_row}", model)
        write_cell(ws, f"{plan1_cols['price']}{excel_row}", price if price else "")
        write_cell(ws, f"{plan1_cols['qty']}{excel_row}", qty if qty else "")
        write_cell(ws, f"{plan1_cols['total']}{excel_row}", total if total else "")

    if not is_single_mode:
        for index, row in enumerate(plan2_rows[: len(config["plan2_rows"])]):
            excel_row = config["plan2_rows"][index]
            plan2_cols = config["plan2_cols"]
            model = row.get("model", "")
            price = parse_number(row.get("price"))
            qty = parse_number(row.get("qty"))
            promo_price = parse_number(row.get("promo_price"))
            total = price * qty
            promo_total = promo_price * qty

            write_model_cell(ws, f"{plan2_cols['model']}{excel_row}", model)
            write_cell(ws, f"{plan2_cols['qty']}{excel_row}", qty if qty else "")

            if plan2_promo_enabled and promo_price:
                promo_row = excel_row + config["plan2_promo_offset"]
                write_cell(ws, f"{plan2_cols['price']}{excel_row}", price if price else "")
                write_cell(ws, f"{plan2_cols['total']}{excel_row}", total if total else "")
                write_cell(ws, f"{plan2_cols['price']}{promo_row}", promo_price if promo_price else "")
                write_cell(ws, f"{plan2_cols['total']}{promo_row}", promo_total if promo_total else "")
                resolve_target_cell(ws, f"{plan2_cols['price']}{promo_row}").font = resolve_target_cell(ws, f"{plan2_cols['price']}{promo_row}").font.copy(
                    color="FF0000"
                )
                resolve_target_cell(ws, f"{plan2_cols['total']}{promo_row}").font = resolve_target_cell(ws, f"{plan2_cols['total']}{promo_row}").font.copy(
                    color="FF0000"
                )
            else:
                promo_row = excel_row + config["plan2_promo_offset"]
                write_cell(ws, f"{plan2_cols['price']}{excel_row}", price if price else "")
                write_cell(ws, f"{plan2_cols['total']}{excel_row}", total if total else "")
                if promo_row != excel_row:
                    write_cell(ws, f"{plan2_cols['price']}{promo_row}", "")
                    write_cell(ws, f"{plan2_cols['total']}{promo_row}", "")

    return wb


def build_excel_bytes(**kwargs):
    wb = fill_template(**kwargs)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_final_pdf(quote_pdf_bytes):
    if not INTRO_PDF_SOURCE.exists():
        raise FileNotFoundError(f"표지 PDF를 찾을 수 없습니다: {INTRO_PDF_SOURCE}")

    intro_reader = PdfReader(BytesIO(INTRO_PDF_SOURCE.read_bytes()))
    quote_reader = PdfReader(BytesIO(quote_pdf_bytes))
    writer = PdfWriter()
    append_debug_log(
        "build_final_pdf "
        f"intro_pages={len(intro_reader.pages)} quote_pages={len(quote_reader.pages)}"
    )

    for page in intro_reader.pages:
        writer.add_page(page)
    for page in quote_reader.pages:
        writer.add_page(page)

    output = BytesIO()
    writer.write(output)
    append_debug_log(
        f"build_final_pdf done final_pages={len(writer.pages)} final_bytes={output.tell()}"
    )
    return output.getvalue()


def expand_product_pdf_links(pdf_bytes, product_models):
    targets = {
        str(model or "").strip(): CHUNGHO_PRODUCT_LINKS[str(model or "").strip()]
        for model in product_models
        if str(model or "").strip() in CHUNGHO_PRODUCT_LINKS
    }
    if not targets:
        return pdf_bytes

    reader = PdfReader(BytesIO(pdf_bytes))
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)

    for page_index, page in enumerate(reader.pages):
        matches = []

        def collect_text(text, cm, tm, font_dict, font_size):
            cleaned = str(text or "").strip()
            if cleaned not in targets:
                return
            x = float(tm[4])
            y = float(tm[5])
            size = max(float(font_size or 0), 8.0)
            matches.append((cleaned, x, y, size))

        page.extract_text(visitor_text=collect_text)
        for model, x, y, size in matches:
            width = max(size * len(model), size * 2.5)
            writer.add_uri(
                page_number=page_index,
                uri=targets[model],
                rect=(x, y - size * 0.25, x + width, y + size * 1.15),
            )

    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def convert_excel_to_pdf(xlsx_bytes):
    started_at = time.perf_counter()
    append_debug_log(f"convert_excel_to_pdf start xlsx_bytes={len(xlsx_bytes)}")
    PDF_WORK_DIR.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="quote_", dir=PDF_WORK_DIR) as tmp_dir_name:
        tmp_dir = Path(tmp_dir_name).resolve()
        xlsx_path = (tmp_dir / "quote.xlsx").resolve()
        pdf_path = (tmp_dir / "quote.pdf").resolve()
        xlsx_path.write_bytes(xlsx_bytes)
        append_debug_log(f"convert_excel_to_pdf tmp_dir={tmp_dir}")

        errors = []
        if platform.system().lower().startswith("win"):
            try:
                append_debug_log("convert_excel_to_pdf excel_worker_start")
                result = subprocess.run(
                    [sys.executable, str(EXCEL_PDF_WORKER), str(xlsx_path), str(pdf_path)],
                    capture_output=True,
                    text=True,
                    timeout=30,
                    check=False,
                )
                if result.returncode == 0 and pdf_path.exists():
                    elapsed = time.perf_counter() - started_at
                    append_debug_log(
                        "convert_excel_to_pdf excel_success "
                        f"elapsed={elapsed:.2f}s pdf_bytes={pdf_path.stat().st_size}"
                    )
                    return pdf_path.read_bytes()
                error_text = result.stderr or result.stdout or "Excel 변환 실패"
                append_debug_log(
                    "convert_excel_to_pdf excel_worker_failed "
                    f"returncode={result.returncode} error={error_text!r}"
                )
                errors.append(error_text)
            except subprocess.TimeoutExpired:
                append_debug_log("convert_excel_to_pdf excel_worker_timeout after 30s")
                errors.append("Excel 변환 시간 초과(30초)")
            except Exception as exc:
                append_debug_log(f"convert_excel_to_pdf excel_exception error={exc}")
                errors.append(f"Excel 변환 실패: {exc}")

        soffice = shutil.which("soffice")
        if soffice:
            append_debug_log(f"convert_excel_to_pdf soffice_start path={soffice}")
            cmd = [
                soffice,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(tmp_dir),
                str(xlsx_path),
            ]
            result = subprocess.run(cmd, capture_output=True, text=True, check=False)
            if result.returncode == 0 and pdf_path.exists():
                elapsed = time.perf_counter() - started_at
                append_debug_log(
                    "convert_excel_to_pdf soffice_success "
                    f"elapsed={elapsed:.2f}s pdf_bytes={pdf_path.stat().st_size}"
                )
                return pdf_path.read_bytes()
            append_debug_log(
                "convert_excel_to_pdf soffice_failed "
                f"returncode={result.returncode} stderr={result.stderr!r} stdout={result.stdout!r}"
            )
            errors.append(result.stderr or result.stdout or "LibreOffice 변환 실패")

        if not errors:
            errors.append("PDF 변환 도구를 찾을 수 없습니다.")
        append_debug_log(f"convert_excel_to_pdf failed errors={errors!r}")
        raise RuntimeError(" / ".join(errors))


def resolve_gmail_credentials():
    if HARDCODE_GMAIL_USER and HARDCODE_GMAIL_APP_PASSWORD:
        return HARDCODE_GMAIL_USER, HARDCODE_GMAIL_APP_PASSWORD
    secrets = getattr(st, "secrets", {})
    user = secrets.get("gmail_user") or os.getenv("GMAIL_USER")
    password = secrets.get("gmail_app_password") or os.getenv("GMAIL_APP_PASSWORD")
    return user, password


def build_email_html(body_text, catalog_links):
    link_items = "".join(
        f'<li><a href="{url}" target="_blank" rel="noopener noreferrer">{label}</a></li>'
        for label, url in catalog_links
    )
    return f"""
    <html>
      <body style="font-family: Arial, sans-serif; line-height: 1.6;">
        <p>{body_text}</p>
        <p>카탈로그 링크:</p>
        <ul>{link_items}</ul>
      </body>
    </html>
    """


def send_email(to_email, subject, body, pdf_bytes, pdf_name):
    smtp_user, smtp_password = resolve_gmail_credentials()
    if not smtp_user or not smtp_password:
        raise RuntimeError(
            "Gmail 계정 정보가 없습니다. st.secrets 또는 환경변수에 설정하세요."
        )

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = smtp_user
    msg["To"] = to_email
    msg.set_content(body)
    msg.add_alternative(build_email_html(body, CATALOG_LINKS), subtype="html")
    msg.add_attachment(
        pdf_bytes,
        maintype="application",
        subtype="pdf",
        filename=pdf_name,
    )

    with smtplib.SMTP("smtp.gmail.com", 587, timeout=30) as server:
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)


def build_filename(recipient, quote_date):
    recipient_label = recipient.strip() if recipient else ""
    base_label = f"{recipient_label}귀하" if recipient_label else "견적서"
    date_label = format_file_date(quote_date)
    return sanitize_filename(f"{base_label}{date_label}")


def save_artifacts(save_dir, file_stem, xlsx_bytes, pdf_bytes=None):
    save_path = Path(save_dir)
    save_path.mkdir(parents=True, exist_ok=True)
    xlsx_path = save_path / f"{file_stem}.xlsx"
    xlsx_path.write_bytes(xlsx_bytes)
    if pdf_bytes:
        pdf_path = save_path / f"{file_stem}.pdf"
        pdf_path.write_bytes(pdf_bytes)


def handle_event(event, save_dir):
    request_started_at = time.perf_counter()
    action = event.get("action")
    request_id = event.get("request_id") or str(int(time.time() * 1000))
    data = event.get("data") or {}
    append_debug_log(f"handle_event start request_id={request_id} action={action}")

    if action not in {"download_pdf", "preview_pdf", "send_email"}:
        return {
            "id": request_id,
            "type": "message",
            "message": "알 수 없는 요청입니다.",
        }

    recipient_raw = str(data.get("recipient") or "").strip()
    recipient = recipient_raw
    ext = str(data.get("ext") or "").strip()
    email = str(data.get("email") or "").strip()
    remarks = str(data.get("remarks") or "").strip()
    if action == "send_email" and not email:
        append_debug_log(f"handle_event send_email_rejected request_id={request_id} reason=missing_email")
        return {
            "id": request_id,
            "type": "message",
            "message": "이메일 주소를 입력하세요.",
        }
    mode = str(data.get("mode") or "compare").strip().lower()
    quote_date = parse_date(data.get("quote_date"))

    plan1_rows = (data.get("plan1") or {}).get("rows") or []
    plan2_rows = (data.get("plan2") or {}).get("rows") or []
    if len(plan1_rows) > MAX_TEMPLATE_ROWS or len(plan2_rows) > MAX_TEMPLATE_ROWS:
        return {
            "id": request_id,
            "type": "message",
            "message": f"제품은 각 견적안에 최대 {MAX_TEMPLATE_ROWS}개까지 입력할 수 있습니다.",
        }
    plan2_promo_enabled = bool((data.get("plan2") or {}).get("promo_enabled"))
    plan1_rows_norm = normalize_rows(plan1_rows)
    plan2_rows_norm = normalize_rows(plan2_rows)

    if mode == "single":
        plan2_rows_norm = [{"model": "", "price": 0, "qty": 0, "promo_price": 0}] * MAX_TEMPLATE_ROWS
        plan2_promo_enabled = False

    plan1_total = compute_totals(plan1_rows_norm)
    plan2_total = compute_totals(plan2_rows_norm, use_promo=plan2_promo_enabled)

    discount1 = max(parse_float(data.get("discount1")), 0.0)
    discount2 = max(parse_float(data.get("discount2")), 0.0)
    plan1_effective = (
        int(round(plan1_total * (1 - discount1))) if discount1 > 0 else plan1_total
    )
    plan2_effective = (
        int(round(plan2_total * (1 - discount2))) if discount2 > 0 else plan2_total
    )
    plan1_monthly = plan1_total
    plan2_monthly = plan2_total
    plan1_prepay = int(round(plan1_total * (1 - discount1))) if discount1 > 0 else ""
    plan2_prepay = int(round(plan2_total * (1 - discount2))) if discount2 > 0 else ""
    summary_text = build_summary_text(plan1_effective, plan2_effective)
    if plan2_promo_enabled:
        summary_text = f"{summary_text} (프로모션적용)"

    xlsx_bytes = build_excel_bytes(
        recipient=recipient,
        ext=ext,
        quote_date=quote_date,
        email=email,
        mode=mode,
        plan1_rows=plan1_rows_norm,
        plan2_rows=plan2_rows_norm,
        plan1_total=plan1_monthly,
        plan2_total=plan2_monthly,
        plan1_prepay=plan1_prepay,
        plan2_prepay=plan2_prepay,
        plan2_promo_enabled=plan2_promo_enabled,
        summary_text=summary_text,
        remarks=remarks,
    )
    append_debug_log(
        "handle_event build_excel_bytes_done "
        f"request_id={request_id} elapsed={time.perf_counter() - request_started_at:.2f}s "
        f"xlsx_bytes={len(xlsx_bytes)} mode={mode} plan1_rows={len(plan1_rows_norm)} plan2_rows={len(plan2_rows_norm)} "
        f"promo={plan2_promo_enabled}"
    )

    file_stem = build_filename(recipient_raw, quote_date)

    try:
        quote_pdf_bytes = convert_excel_to_pdf(xlsx_bytes)
        product_models = [row.get("model", "") for row in plan1_rows_norm]
        if mode != "single":
            product_models.extend(row.get("model", "") for row in plan2_rows_norm)
        quote_pdf_bytes = expand_product_pdf_links(quote_pdf_bytes, product_models)
        append_debug_log(
            "handle_event excel_pdf_done "
            f"request_id={request_id} elapsed={time.perf_counter() - request_started_at:.2f}s "
            f"quote_pdf_bytes={len(quote_pdf_bytes)}"
        )
        pdf_bytes = build_final_pdf(quote_pdf_bytes)
        append_debug_log(
            "handle_event final_pdf_done "
            f"request_id={request_id} elapsed={time.perf_counter() - request_started_at:.2f}s "
            f"final_pdf_bytes={len(pdf_bytes)}"
        )
    except Exception as exc:
        append_debug_log(
            "handle_event exception "
            f"request_id={request_id} elapsed={time.perf_counter() - request_started_at:.2f}s "
            f"error={exc}"
        )
        return {
            "id": request_id,
            "type": "message",
            "message": f"PDF 생성 실패: {exc}",
        }

    if action == "send_email":
        try:
            doc_label = "단일견적서" if mode == "single" else "비교견적서"
            if recipient_raw:
                subject = f"{recipient_raw} 귀하 {doc_label} 검토 부탁드리겠습니다"
            else:
                subject = f"{doc_label} 검토 부탁드리겠습니다"
            body = "첨부된 PDF 견적서를 확인해 주세요."
            send_email(
                to_email=email,
                subject=subject,
                body=body,
                pdf_bytes=pdf_bytes,
                pdf_name=f"{file_stem}.pdf",
            )
            append_debug_log(
                f"handle_event send_email_success request_id={request_id} to={email}"
            )
            save_artifacts(save_dir, file_stem, xlsx_bytes=xlsx_bytes, pdf_bytes=pdf_bytes)
            return {
                "id": request_id,
                "type": "success",
                "message": "이메일 전송 완료",
            }
        except Exception as exc:
            append_debug_log(
                f"handle_event send_email_failed request_id={request_id} error={exc}"
            )
            return {
                "id": request_id,
                "type": "message",
                "message": f"전송 실패: {exc}",
            }

    if action == "download_pdf":
        return {
            "id": request_id,
            "type": "pdf",
            "filename": f"{file_stem}.pdf",
            "content": base64.b64encode(pdf_bytes).decode("ascii"),
            "message": "PDF 생성 완료",
        }

    if action == "preview_pdf":
        return {
            "id": request_id,
            "type": "preview",
            "filename": f"{file_stem}.pdf",
            "content": base64.b64encode(pdf_bytes).decode("ascii"),
        }

    return {
        "id": request_id,
        "type": "message",
        "message": "알 수 없는 요청입니다.",
    }


def safe_rerun():
    if hasattr(st, "rerun"):
        st.rerun()
        return
    if hasattr(st, "experimental_rerun"):
        st.experimental_rerun()


st.set_page_config(
    page_title="가격 비교",
    layout="wide",
    initial_sidebar_state="collapsed",
)

DEFAULT_SAVE_DIR.mkdir(parents=True, exist_ok=True)
settings = load_settings()
if "save_dir" not in st.session_state:
    st.session_state["save_dir"] = settings.get("save_dir", str(DEFAULT_SAVE_DIR))

with st.sidebar:
    st.markdown("저장 설정")
    save_dir_input = st.text_input(
        "저장 폴더",
        value=st.session_state["save_dir"],
    )
    if st.button("저장 위치 저장"):
        try:
            saved_dir = persist_save_dir(save_dir_input)
            st.session_state["save_dir"] = saved_dir
            st.success("저장 완료")
        except Exception as exc:
            st.error(f"저장 실패: {exc}")
    st.caption(f"현재 위치: {st.session_state['save_dir']}")
    st.divider()
    st.caption("Gmail 앱 비밀번호는 st.secrets 또는 환경변수로 설정하세요.")

st.markdown(
    """
    <style>
      #MainMenu {visibility: hidden;}
      footer {visibility: hidden;}
      header {visibility: hidden;}
      .block-container {padding: 0;}
    </style>
    """,
    unsafe_allow_html=True,
)

if not HTML_PATH.exists():
    st.error(f"HTML 파일을 찾을 수 없습니다: {HTML_PATH}")
    st.stop()

if not COMPONENT_DIR.exists():
    st.error(f"컴포넌트 폴더를 찾을 수 없습니다: {COMPONENT_DIR}")
    st.stop()

html = HTML_PATH.read_text(encoding="utf-8")
html = inject_responsive_layout(html)
html = inject_catalog_data(html)
html = inject_product_info(html)
html = inject_initial_view(html, st.session_state.get("active_view"))
html = inject_bridge(html)
html_hash = hashlib.sha256(html.encode("utf-8")).hexdigest()

price_compare_component = components.declare_component(
    "price_compare",
    path=str(COMPONENT_DIR),
)

response = st.session_state.get("bridge_response")
last_request_id = st.session_state.get("last_request_id")

payload = price_compare_component(
    html=html,
    html_hash=html_hash,
    response=response,
    height=COMPONENT_HEIGHT,
)

if payload and isinstance(payload, dict):
    if payload.get("action") == "ack_response":
        response_id = payload.get("response_id")
        current_response = st.session_state.get("bridge_response") or {}
        if current_response.get("id") == response_id:
            st.session_state.pop("bridge_response", None)
            append_debug_log(f"bridge response_ack response_id={response_id}")
        st.stop()

    request_id = payload.get("request_id")
    if request_id and request_id != last_request_id:
        append_debug_log(
            f"bridge request_received request_id={request_id} action={payload.get('action')}"
        )
        st.session_state["last_request_id"] = request_id
        view = (payload.get("data") or {}).get("view")
        if view:
            st.session_state["active_view"] = view
        response = handle_event(
            payload,
            st.session_state["save_dir"],
        )
        st.session_state["bridge_response"] = response
        safe_rerun()
